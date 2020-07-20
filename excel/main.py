# This code required Third party modlue
#	following code for line 1 to 7
# 	Check if required module is present in system or not...
from Dependencies import check_module
check_module({'openpyxl'})
# Module Checking end...
# Remove This part if not required.... :)

import openpyxl as xl
from openpyxl.chart import PieChart, Reference
import openpyxl.styles as style

# Load xlsx book
wb = xl.load_workbook("./excel/transactions.xlsx")



# # #################### WORKING WITH CELL AND DATA ############
# Select sheet
sheet = wb["Sheet1"]


# Select cell
cell1 = sheet["a1"]
cell2 = sheet.cell(1 , 1)
print(cell1.value, cell2.value)


# Max row & Column
print(sheet.max_row, sheet.max_column)


# Update Value
for row in range( 1, 5 ):
	price_cell = sheet.cell(row, 3)
	if row == 1:
		sheet.cell(1, 4).value = "Discount Amount"
	else:
		sheet.cell(row, 4).value = price_cell.value*0.9


# Read all from sheet
for row in sheet:
	for cell in row:
		print("%-15s" %cell.value, end="")
	print()


# Get cell detail
cell = sheet["C3"]
print(f"Row no:{cell.row}\nColumn no.:{cell.column}\nCell Co-ordinate:{cell.coordinate}")


# Read a perticular row or column
row = sheet["1"]
col = sheet["1"]
print(row)
print(col)


# Read in range
range_block = sheet["b2:c3"]
print(range_block)


# Add date using list and tuple
new_data = [
	("Data11", "Data12", 13),
	("Data21", "Data22", 23) ]

for data in new_data:
	sheet.append(data)


# Desiging cells
for cell in sheet["1"]:
	cell.font = style.Font(bold=True, color="ffffff")
	cell.fill = style.PatternFill(start_color="0000ff", fill_type="solid")


# Plot Bar chart
chart_values = xl.chart.Reference ( sheet,
									min_row=2, max_row=4,
									min_col=4, max_col=4 )

chart_type = xl.chart.BarChart()
chart_type.add_data( chart_values )
sheet.add_chart( chart_type, "e1" )


# Plot pie char with details
chart_title = "Pie Chart"
chart_catagories = xl.chart.Reference(sheet,
								min_col=1,
								min_row=2, max_row=4 )
chart_data = Reference(sheet,
							min_col=4,
							min_row=1, max_row=4 )

chart = PieChart()
chart.title = chart_title
chart.set_categories(chart_catagories)
chart.add_data(chart_data, titles_from_data=True)

sheet.add_chart(chart, "e1")


# Insert row/cols from index to length
sheet.insert_rows( 2, 2 )
sheet.insert_cols( 2, 2 )


# Delete rows/coloumn from index to length
sheet.delete_rows( 5, 3 )
sheet.delete_cols( 5, 2 )


wb.save("./excel/transactions.xlsx")


# #################### WORKING WITH SHEETS ############

# Get sheets list
print(wb.sheetnames)	# Return name stirng
print(wb.worksheets)	# Return Worksheet object

# Create new Sheet
wb.create_sheet("Sheet2")
wb.create_sheet("Sheet0", 0)
print(wb.sheetnames)	# Return name stirng

# Delete worksheet
wb.remove_sheet(wb["Sheet0"])
wb.remove_sheet(wb["Sheet2"])
print(wb.sheetnames)	# Return name stirng

wb.save("./excel/transactions.xlsx")
